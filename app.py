from flask import Flask, render_template, request
from openpyxl import load_workbook
import openai

app = Flask(__name__)

# Função para verificar se o email já existe no arquivo Excel
def check_existing_email(email):
    try:
        wb = load_workbook('user_data.xlsx')
        ws = wb.active
        existing_emails = [row[0].value for row in ws.iter_rows(min_row=2, max_col=1)]
        return email in existing_emails
    except FileNotFoundError:
        return False  # Se o arquivo não existe, o email não pode existir

# Função para salvar o email, senha e nome de usuário em um arquivo Excel local
def save_credentials_to_excel(username, email, password):
    if not check_existing_email(email):
        try:
            wb = load_workbook('user_data.xlsx')
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.append(['Username', 'Email', 'Password'])
        ws.append([username, email, password])
        wb.save('user_data.xlsx')
        return True  # Retorna True se os dados forem salvos com sucesso
    else:
        return False  # Retorna False se o email já existir


@app.route('/logado', methods=['POST'])
def logado():
    if request.method == 'POST':
        username = request.form.get("username")
        email = request.form.get("email")
        password = request.form.get("password")
        if save_credentials_to_excel(username, email, password):
            return render_template('logged_in.html', email=email)
        else:
            return render_template('registration_failed.html', message="Email already exists")
    else:
        return "Method Not Allowed", 405
    
@app.route('/')
def index():
    return render_template('login.html')

@app.route('/chatbot')
def chatbot():
    return render_template('chatbot.html')

if __name__ == '__main__':
    app.run(debug=True)
